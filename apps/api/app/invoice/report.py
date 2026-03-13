from __future__ import annotations

import csv
import io
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .schemas import ValidatedInvoice

_INVOICE_HEADERS = [
    "File", "Vendor", "Vendor (Original)", "Invoice #", "Date", "Due Date",
    "Currency", "Subtotal", "Tax", "Total", "Language", "Confidence", "Notes",
]

_FLAG_HEADERS = [
    "Invoice #", "File", "Vendor", "Type", "Severity", "Message",
]


async def generate_excel_report(invoices: list[ValidatedInvoice]) -> bytes:
    """Generate an Excel report with Invoices and Flags sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # --- Sheet 1: Invoices ---
    ws_inv = wb.active
    ws_inv.title = "Invoices"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col_idx, header in enumerate(_INVOICE_HEADERS, 1):
        cell = ws_inv.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, vi in enumerate(invoices, 2):
        inv = vi.invoice
        ws_inv.cell(row=row_idx, column=1, value=inv.source_filename)
        ws_inv.cell(row=row_idx, column=2, value=inv.vendor_name)
        ws_inv.cell(row=row_idx, column=3, value=inv.vendor_name_original)
        ws_inv.cell(row=row_idx, column=4, value=inv.invoice_number)
        ws_inv.cell(row=row_idx, column=5, value=inv.invoice_date)
        ws_inv.cell(row=row_idx, column=6, value=inv.due_date)
        ws_inv.cell(row=row_idx, column=7, value=inv.currency)
        ws_inv.cell(row=row_idx, column=8, value=inv.subtotal)
        ws_inv.cell(row=row_idx, column=9, value=inv.tax)
        ws_inv.cell(row=row_idx, column=10, value=inv.total)
        ws_inv.cell(row=row_idx, column=11, value=inv.language)
        ws_inv.cell(row=row_idx, column=12, value=inv.confidence)
        ws_inv.cell(row=row_idx, column=13, value=vi.accountant_notes)

    # Auto-width columns
    for col in ws_inv.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_inv.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # --- Sheet 2: Flags ---
    ws_flags = wb.create_sheet("Flags")

    for col_idx, header in enumerate(_FLAG_HEADERS, 1):
        cell = ws_flags.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    severity_fills = {
        "critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }

    flag_row = 2
    for vi in invoices:
        inv = vi.invoice
        for flag in vi.flags:
            ws_flags.cell(row=flag_row, column=1, value=flag.invoice_index)
            ws_flags.cell(row=flag_row, column=2, value=inv.source_filename)
            ws_flags.cell(row=flag_row, column=3, value=inv.vendor_name)
            ws_flags.cell(row=flag_row, column=4, value=flag.type)
            sev_cell = ws_flags.cell(row=flag_row, column=5, value=flag.severity)
            if flag.severity in severity_fills:
                sev_cell.fill = severity_fills[flag.severity]
            ws_flags.cell(row=flag_row, column=6, value=flag.message)
            flag_row += 1

    for col in ws_flags.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_flags.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_csv_report(invoices: list[ValidatedInvoice]) -> bytes:
    """Generate a CSV report (invoices only, no flags sheet)."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_INVOICE_HEADERS)

    for vi in invoices:
        inv = vi.invoice
        writer.writerow([
            inv.source_filename,
            inv.vendor_name,
            inv.vendor_name_original or "",
            inv.invoice_number or "",
            inv.invoice_date or "",
            inv.due_date or "",
            inv.currency,
            inv.subtotal,
            inv.tax,
            inv.total,
            inv.language,
            inv.confidence,
            vi.accountant_notes or "",
        ])

    return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
